from bs4 import BeautifulSoup
import openpyxl
import os

def processar_arquivo_html(caminho_do_arquivo, sheet):
    with open(caminho_do_arquivo, 'r', encoding='utf-8') as file:
        html_content = file.read()

    soup = BeautifulSoup(html_content, 'html.parser')

    # Encontrar a tabela de Request Statistics
    request_table = soup.find('div', class_='requests').find('table', class_='stats')

    if request_table:
        # Encontrar as linhas da tabela
        rows = request_table.find('tbody').find_all('tr')

        # Inicializar variável para subtrair os valores
        result_requests_fails = []

        # Adicionar dados das linhas à planilha e subtrair os valores
        for row_index, row in enumerate(rows, start=sheet.max_row + 2):  # Adiciona duas linhas em branco
            data = [cell.text.strip() for cell in row.find_all('td')]
            num_requests, num_fails = int(data[2]), int(data[3])
            result = num_requests - num_fails
            result_requests_fails.append((result, num_fails))
            #sheet.cell(row=row_index, column=1, value=num_requests)
            sheet.cell(row=row_index, column=1, value=result)
            sheet.cell(row=row_index, column=2, value=num_fails)

        """ # Imprimir resultado no console
        print(f"\nResultados de {caminho_do_arquivo} (Requests - Fails)\t# Requests\t# Fails")
        for result, num_fails in result_requests_fails:
            print(f'{result}\t{num_requests}\t{num_fails}') """
    else:
        print(f"Tabela não encontrada em {caminho_do_arquivo}")

# Criar um novo arquivo Excel
workbook = openpyxl.Workbook()
consolidated_sheet = workbook.active
consolidated_sheet.title = "Consolidado"
consolidated_sheet.append([ 'Resultado (Requests - Fails)', '# Fails'])

# Substitua pelos caminhos reais dos seus arquivos HTML
caminhos_arquivos_html = [
    '/home/kaique/Documentos/ufg/MEC_RNIS/Teste/time_10m_client_1_2000.html',
    '/home/kaique/Documentos/ufg/MEC_RNIS/Teste/time_10m_client_2_2000.html',
    '/home/kaique/Documentos/ufg/MEC_RNIS/Teste/time_10m_client_3_2000.html',
    '/home/kaique/Documentos/ufg/MEC_RNIS/Teste/time_10m_client_4_2000.html',
    '/home/kaique/Documentos/ufg/MEC_RNIS/Teste/time_10m_client_5_2000.html',
    '/home/kaique/Documentos/ufg/MEC_RNIS/Teste/time_10m_client_6_2000.html',
    '/home/kaique/Documentos/ufg/MEC_RNIS/Teste/time_10m_client_7_2000.html',
    '/home/kaique/Documentos/ufg/MEC_RNIS/Teste/time_10m_client_8_2000.html',
    '/home/kaique/Documentos/ufg/MEC_RNIS/Teste/time_10m_client_9_2000.html',
    '/home/kaique/Documentos/ufg/MEC_RNIS/Teste/time_10m_client_10_2000.html'
]

for caminho_arquivo_html in caminhos_arquivos_html:
    if os.path.exists(caminho_arquivo_html):
        processar_arquivo_html(caminho_arquivo_html, consolidated_sheet)
    else:
        print(f"Arquivo não encontrado: {caminho_arquivo_html}")

# Ajustar largura das colunas em todas as planilhas
for sheet in workbook.sheetnames:
    current_sheet = workbook[sheet]
    for column in current_sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        current_sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = adjusted_width

# Salvar a planilha consolidada em um arquivo
planilha_caminho = '/home/kaique/Documentos/ufg/MEC_RNIS/Teste/planilha_consolidada.xlsx'  # Substitua pelo caminho desejado
workbook.save(planilha_caminho)
print(f"\nPlanilha consolidada salva em: {planilha_caminho}")
